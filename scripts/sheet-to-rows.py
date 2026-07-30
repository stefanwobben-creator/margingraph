#!/usr/bin/env python3
"""Turn a spreadsheet into the grid the intake reader works on.

Everything clever happens in `lib/reports/intake`, which is TypeScript, tested,
and shipped. This is the dumbest possible bridge to it: open the file, read the
cells, write JSON. It deliberately makes no decisions about which column is the
budget or which row is turnover, because a decision made here would be a
decision nobody can test.

    python3 scripts/sheet-to-rows.py "Q1 2026.xlsx" > /tmp/rows.json
    python3 scripts/sheet-to-rows.py accounts.csv --sheet 2

Reads .xlsx (openpyxl), .csv, and .pdf (pdfplumber). Formula cells come back
as their last cached value, which is what the sender saw on their screen; a
workbook saved by something that does not cache values will show blanks there,
and the reader will say so rather than inventing figures.

A PDF is read by finding the page that calls itself the profit and loss
account and turning each printed line into label-plus-numbers. A scanned PDF
has no text to find, and the script says that instead of returning an empty
grid that would fail somewhere less explicable.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path


def from_csv(path: Path) -> dict:
    """A CSV has no types, so anything that parses as a number becomes one."""
    with path.open(newline="", encoding="utf-8-sig") as handle:
        sample = handle.read(8192)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        rows = [[cell_value(c) for c in row] for row in csv.reader(handle, dialect)]
    return {"name": path.stem, "rows": rows}


def cell_value(raw: str):
    text = raw.strip()
    if text == "":
        return None
    # European and English notation, plus the accountant's parenthesised
    # negative, which is a minus sign written as punctuation.
    negative = text.startswith("(") and text.endswith(")")
    body = text[1:-1] if negative else text
    body = body.replace("€", "").replace("%", "").strip()
    cleaned = body.replace(" ", "").replace(" ", "")
    if "," in cleaned and "." in cleaned:
        cleaned = (
            cleaned.replace(".", "").replace(",", ".")
            if cleaned.rfind(",") > cleaned.rfind(".")
            else cleaned.replace(",", "")
        )
    elif "," in cleaned:
        # "2,987,843" is thousands; "843,50" is a decimal. Groups of three
        # after the first say which one this is.
        cleaned = (
            cleaned.replace(",", "")
            if re.fullmatch(r"-?\d{1,3}(,\d{3})+", cleaned)
            else cleaned.replace(",", ".")
        )
    elif "." in cleaned:
        # "3.861.609" is how a Dutch statement writes three million. Read as
        # a decimal it becomes three euros and a rounding error, which is the
        # single worst misreading this bridge can produce.
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})+", cleaned):
            cleaned = cleaned.replace(".", "")
    try:
        number = float(cleaned)
    except ValueError:
        return text
    if negative:
        number = -number
    return -0.0 + number if number else 0.0


def from_xlsx(path: Path, sheet: str | None) -> dict:
    try:
        from openpyxl import load_workbook
    except ImportError:  # pragma: no cover - environment problem, not logic
        sys.exit("openpyxl is not installed. Run: pip install openpyxl")

    book = load_workbook(path, data_only=True)
    if sheet is not None:
        try:
            worksheet = book[sheet] if sheet in book.sheetnames else book.worksheets[int(sheet) - 1]
        except (ValueError, IndexError):
            sys.exit(f"no sheet {sheet!r}. This file has: {', '.join(book.sheetnames)}")
    else:
        # The first sheet with anything on it. A workbook that opens on an
        # empty cover tab is common enough to be worth handling.
        worksheet = next(
            (ws for ws in book.worksheets if ws.max_row and ws.max_row > 1),
            book.worksheets[0],
        )

    # Accounting exports write a group header once, merged across the columns
    # it covers: "Budget huidige jaar" sits in one cell above five columns and
    # every other cell in the range is empty. Read literally, four of those five
    # columns lose the only word that says what they are, and two different
    # years both end up called "Period actual". Copying the value across the
    # range is not interpretation, it is what the file already shows on screen.
    #
    # Horizontal merges only. A vertical merge means one label spanning several
    # rows, and repeating it would invent lines that are not there.
    for merged in list(worksheet.merged_cells.ranges):
        if merged.min_row != merged.max_row:
            continue
        value = worksheet.cell(merged.min_row, merged.min_col).value
        if value is None:
            continue
        worksheet.unmerge_cells(str(merged))
        for column in range(merged.min_col, merged.max_col + 1):
            worksheet.cell(merged.min_row, column).value = value

    rows = []
    for row in worksheet.iter_rows(values_only=True):
        rows.append(
            [
                cell if isinstance(cell, (int, float)) and not isinstance(cell, bool)
                else (str(cell).strip() or None) if cell is not None
                else None
                for cell in row
            ]
        )
    return {"name": worksheet.title, "rows": rows, "sheets": book.sheetnames}


# The names a profit and loss account goes by on its own page. The negative
# check matters: "Toelichting op de winst-en-verliesrekening" introduces the
# detail pages, and reading those as well would count every line twice.
PNL_HEADING = re.compile(
    r"winst[\s-]*en[\s-]*verliesrekening|profit\s+and\s+loss|staat\s+van\s+baten\s+en\s+lasten|income\s+statement",
    re.IGNORECASE,
)
PNL_DETAIL = re.compile(r"toelichting", re.IGNORECASE)


def pdf_number(token: str):
    """A numeric token as printed in accounts, or None if it is not one.

    Dutch statements write thousands with dots and sometimes hang the minus
    after the number. Parentheses are a minus written as punctuation. A year
    is a number too, which is fine: the reader treats a header row of years
    as headers because they sit above the first label-with-figures row.
    """
    text = token.strip()
    trailing_minus = text.endswith("-") and len(text) > 1
    if trailing_minus:
        text = text[:-1]
    if not re.fullmatch(r"\(?-?€?\s*[\d.,]+\)?", text):
        return None
    value = cell_value(text)
    if not isinstance(value, (int, float)):
        return None
    return -value if trailing_minus and value > 0 else value


def from_pdf(path: Path) -> dict:
    try:
        import pdfplumber
    except ImportError:  # pragma: no cover - environment problem, not logic
        sys.exit("pdfplumber is not installed. Run: pip install pdfplumber")

    with pdfplumber.open(path) as book:
        texts = [(page.extract_text() or "") for page in book.pages]

        if not any(text.strip() for text in texts):
            sys.exit(
                f"{path.name} has no text layer, so it is a scan or a photo. "
                "We cannot read figures out of pixels without guessing. Send the "
                "spreadsheet, or a PDF exported from the accounting package."
            )

        # The first page that is the statement itself, not the notes on it.
        page_index = next(
            (
                i
                for i, text in enumerate(texts)
                if PNL_HEADING.search(text)
                and not PNL_DETAIL.search(text.splitlines()[0] if text.splitlines() else "")
                and not any(
                    PNL_DETAIL.search(line) and PNL_HEADING.search(line)
                    for line in text.splitlines()[:4]
                )
            ),
            None,
        )
        if page_index is None:
            pages_with_text = sum(1 for t in texts if t.strip())
            sys.exit(
                f"No page in {path.name} calls itself a profit and loss account "
                f"(read {pages_with_text} pages of text). If the statement is in "
                "there under another name, send the spreadsheet instead."
            )

        page = book.pages[page_index]
        words = page.extract_words(use_text_flow=False)

    # Words into printed lines: same baseline, left to right.
    lines: dict[float, list] = {}
    for word in sorted(words, key=lambda w: (round(w["top"] / 3) * 3, w["x0"])):
        lines.setdefault(round(word["top"] / 3) * 3, []).append(word)

    # First pass: per line, split the right-hand numeric tail off the label,
    # keeping each number's right edge. A printed statement is a table drawn
    # with alignment instead of cells, and the right edge is the alignment.
    parsed = []
    for _, tokens in sorted(lines.items()):
        remaining = list(tokens)
        values = []  # (right edge, value)
        while remaining and (value := pdf_number(remaining[-1]["text"])) is not None:
            values.insert(0, (remaining[-1]["x1"], value))
            remaining.pop()
        label = " ".join(t["text"] for t in remaining).strip() or None
        if label is None and not values:
            continue
        # "Winst-en-verliesrekening over 2022" ends in a number that is part
        # of the title, not a figure. Left as a figure it becomes the first
        # data row, and the real year headers below it stop being headers.
        if (
            label
            and PNL_HEADING.search(label)
            and values
            and all(1990 <= v <= 2100 and float(v).is_integer() for _, v in values)
        ):
            label = f"{label} {' '.join(str(int(v)) for _, v in values)}"
            values = []
        parsed.append((label, values))

    # Cluster the right edges into columns, so a line that only prints one
    # figure still puts it under the year it belongs to. Without this, a
    # subtotal with a single amount would land in the first numeric column
    # and quietly become last year's figure.
    edges = sorted({edge for _, values in parsed for edge, _ in values})
    columns: list[float] = []
    for edge in edges:
        if not columns or edge - columns[-1] > 18:
            columns.append(edge)
        else:
            columns[-1] = edge  # widen the cluster to its rightmost member
    place = lambda edge: min(range(len(columns)), key=lambda i: abs(columns[i] - edge))

    grid = []
    for label, values in parsed:
        row = [label] + [None] * len(columns)
        for edge, value in values:
            row[1 + place(edge)] = value
        grid.append(row)

    # A column of note references — small integers pointing into the
    # toelichting — is layout, not figures, and the reader would otherwise
    # have to ask which column is the period. Dropped only when another
    # column plainly holds the amounts.
    def is_note_column(index: int) -> bool:
        cells = [row[index] for row in grid if row[index] is not None]
        return (
            len(cells) > 0
            and all(isinstance(c, (int, float)) and float(c).is_integer() and abs(c) < 100 for c in cells)
            and any(
                any(isinstance(row[other], (int, float)) and abs(row[other]) >= 1000 for row in grid)
                for other in range(1, len(columns) + 1)
                if other != index
            )
        )

    keep = [0] + [i for i in range(1, len(columns) + 1) if not is_note_column(i)]
    grid = [[row[i] for i in keep] for row in grid]

    # A header line of bare years ("2022  2021") arrives as numbers, and the
    # reader only reads text as a header. What is printed as a column heading
    # should reach the reader as one. Only lines above the first real data
    # line qualify — a year amid the figures is a figure.
    first_data = next(
        (
            i
            for i, row in enumerate(grid)
            if row[0] is not None and any(isinstance(c, (int, float)) for c in row[1:])
        ),
        len(grid),
    )
    for row in grid[:first_data]:
        if all(
            isinstance(c, (int, float)) and 1990 <= c <= 2100 and float(c).is_integer()
            for c in row[1:]
            if c is not None
        ) and any(c is not None for c in row[1:]):
            for i in range(1, len(row)):
                if row[i] is not None:
                    row[i] = str(int(row[i]))

    return {
        "name": f"page {page_index + 1} of {path.name}",
        "rows": grid,
        "sheets": [f"page {i + 1}" for i, text in enumerate(texts) if text.strip()],
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("file")
    parser.add_argument("--sheet", help="sheet name, or 1-based index")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        sys.exit(f"no such file: {path}")

    suffix = path.suffix.lower()
    if suffix == ".csv":
        data = from_csv(path)
    elif suffix == ".pdf":
        data = from_pdf(path)
    else:
        data = from_xlsx(path, args.sheet)
    json.dump(data, sys.stdout, ensure_ascii=False, indent=1)
    sys.stdout.write("\n")


if __name__ == "__main__":
    main()
